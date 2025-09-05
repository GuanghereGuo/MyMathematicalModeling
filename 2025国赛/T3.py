import numpy as np
from scipy.optimize import differential_evolution
import openpyxl
import time
import os

# 物理和环境参数
P_M1_0 = np.array([20000.0, 0.0, 2000.0])  # 导弹初始位置
v_M = 300.0  # 导弹速度
P_FY1_0 = np.array([17800.0, 0.0, 1800.0])  # 无人机1初始位置
R_C = 10.0  # 烟幕云团半径
v_sink = 3.0  # 烟雾下降速度
g = 9.8  # 重力加速度

# 真目标关键点列表 (8点外接盒模型)
key_points = [
    np.array([7, 207.0, 0.0]),
    np.array([7, 193.0, 0.0]),
    np.array([-7, 207.0, 0.0]),
    np.array([-7, 193.0, 0.0]),
    np.array([7, 207.0, 10.0]),
    np.array([7, 193.0, 10.0]),
    np.array([-7, 207.0, 10.0]),
    np.array([-7, 193.0, 10.0]),
    # np.array([0.0, 200.0, 0]),
]

# 导弹轨迹预计算
d_M1 = (np.array([0.0, 0.0, 0.0]) - P_M1_0) / np.linalg.norm(
    np.array([0.0, 0.0, 0.0]) - P_M1_0
)


def P_M1(t):
    """计算t时刻导弹M1的位置"""
    return P_M1_0 + v_M * t * d_M1


# 几何计算函数
def dist_point_to_segment(p, a, b):
    """计算点p到线段ab的最短距离"""
    v = b - a
    w = p - a
    c1 = np.dot(w, v)
    if c1 <= 0:
        return np.linalg.norm(p - a)
    c2 = np.dot(v, v)
    if c2 <= c1:
        return np.linalg.norm(p - b)
    b_proj = c1 / c2
    pb = a + b_proj * v
    return np.linalg.norm(p - pb)


def get_shielding_intervals(v_FY1, theta_rad, t_drop, t_fuse, P_FY1_0, dt=1e-2):
    d_FY1 = np.array([np.cos(theta_rad), np.sin(theta_rad), 0.0])
    P_drop = P_FY1_0 + v_FY1 * t_drop * d_FY1
    v_drop = v_FY1 * d_FY1
    P_detonate = P_drop + v_drop * t_fuse + np.array([0, 0, -0.5 * g * t_fuse**2])
    t_detonate = t_drop + t_fuse

    def P_C(t):
        return P_detonate + np.array([0, 0, -v_sink * (t - t_detonate)])

    shielded_times = []
    t_start = t_detonate
    t_end = t_detonate + 20.0

    for t in np.arange(t_start, t_end, dt):
        p_m1_t = P_M1(t)
        p_c_t = P_C(t)
        is_fully_shielded = True
        for p_key in key_points:
            if dist_point_to_segment(p_c_t, p_m1_t, p_key) > R_C:
                is_fully_shielded = False
                break
        if is_fully_shielded:
            shielded_times.append(t)

    if not shielded_times:
        return []

    intervals = []
    start_of_interval = shielded_times[0]
    for i in range(1, len(shielded_times)):
        if shielded_times[i] - shielded_times[i - 1] > dt * 1.5:
            intervals.append((start_of_interval, shielded_times[i - 1]))
            start_of_interval = shielded_times[i]
    intervals.append((start_of_interval, shielded_times[-1]))

    return intervals


def calculate_union_length(intervals):
    if not intervals:
        return 0

    intervals.sort(key=lambda x: x[0])

    merged = []
    if not intervals:
        return 0

    current_start, current_end = intervals[0]

    for next_start, next_end in intervals[1:]:
        if next_start <= current_end:
            current_end = max(current_end, next_end)
        else:
            merged.append((current_start, current_end))
            current_start, current_end = next_start, next_end

    merged.append((current_start, current_end))

    total_length = sum(end - start for start, end in merged)
    return total_length


def objective_function_Q3(params):
    """问题三的目标函数，使用变量变换法处理约束"""
    v_FY1, theta_rad, t_drop1, t_fuse1, dt_drop2, t_fuse2, dt_drop3, t_fuse3 = params

    t_drop2 = t_drop1 + dt_drop2
    t_drop3 = t_drop2 + dt_drop3

    all_intervals = []

    all_intervals.extend(
        get_shielding_intervals(v_FY1, theta_rad, t_drop1, t_fuse1, P_FY1_0)
    )
    all_intervals.extend(
        get_shielding_intervals(v_FY1, theta_rad, t_drop2, t_fuse2, P_FY1_0)
    )
    all_intervals.extend(
        get_shielding_intervals(v_FY1, theta_rad, t_drop3, t_fuse3, P_FY1_0)
    )

    total_length = calculate_union_length(all_intervals)

    return -total_length


def run_optimization_Q3():
    """执行问题三的完整优化流程"""
    print("--- 问题三：单无人机三弹药优化 ---")

    bounds = [
        (70, 140),  # v_FY1: 速度
        (np.pi * 170 / 180, np.pi * 190 / 180),  # θ_FY1: 角度 [170°, 190°]
        (0, 5),  # t_drop1: 投放时间
        (0, 5),  # t_fuse1: 引信时间
        (1, 5),  # dt_drop2: 第二枚弹药相对第一枚的投放时间差
        (0, 5),  # t_fuse2: 第二枚弹药引信时间
        (1, 5),  # dt_drop3: 第三枚弹药相对第二枚的投放时间差
        (0, 5),  # t_fuse3: 第三枚弹药引信时间
    ]

    print("开始优化 (8维空间，计算量较大，请耐心等待)...")
    start_time = time.time()

    result = differential_evolution(
        objective_function_Q3,
        bounds,
        strategy="best1bin",
        maxiter=300,
        popsize=25,
        tol=0.01,
        disp=True,
        workers=-1,
    )

    end_time = time.time()
    print(f"\n优化完成，耗时: {end_time - start_time:.2f} 秒")

    if result.success:
        p = result.x
        max_time = -result.fun

        t_drop1 = p[2]
        t_drop2 = p[2] + p[4]
        t_drop3 = t_drop2 + p[6]

        v_opt, theta_opt = p[0], p[1]
        params_list = [(t_drop1, p[3]), (t_drop2, p[5]), (t_drop3, p[7])]

        print("\n--- 最优策略 (问题三) ---")
        print(f"无人机飞行速度: {v_opt:.2f} m/s")
        print(f"无人机飞行方向: {np.rad2deg(theta_opt):.2f}°")
        print("------------------------------------")
        for i, (td, tf) in enumerate(params_list):
            print(f"弹药{i+1}: 投放时间={td:.2f}s, 引信={tf:.2f}s")
        print("------------------------------------")
        print(f"最大总有效遮蔽时长: {max_time:.4f} s")

        save_results_to_excel_custom_format(v_opt, theta_opt, params_list, max_time)

    else:
        print("\n优化未成功收敛。可以尝试增加 maxiter 或 popsize。")


def save_results_to_excel_custom_format(v, theta_rad, params_list, max_time):
    """将最优策略的详细信息按照指定的报告格式保存到 result1.xlsx"""
    print("\n正在生成 result1.xlsx 文件...")

    # --- 1. 预计算所有需要填入的数据 ---
    direction_deg = np.rad2deg(theta_rad) % 360  # 确保角度在0-360
    d_FY1 = np.array([np.cos(theta_rad), np.sin(theta_rad), 0.0])

    grenade_data = []
    for i, (t_drop, t_fuse) in enumerate(params_list):
        p_drop = P_FY1_0 + v * t_drop * d_FY1
        v_drop = v * d_FY1
        p_detonate = p_drop + v_drop * t_fuse + np.array([0, 0, -0.5 * g * t_fuse**2])
        grenade_data.append(
            {"id": f"G{i+1}", "p_drop": p_drop, "p_detonate": p_detonate}
        )

    # --- 2. 使用 openpyxl 创建并写入工作簿 ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "问题三最优策略"

    # 写入标题行
    ws.append(["", "1", "2", "3"])

    # 写入数据行
    ws.append(["无人机运动方向", direction_deg, direction_deg, direction_deg])
    ws.append(["无人机运动速度 (m/s)", v, v, v])
    ws.append(
        [
            "烟幕干扰弹编号",
            grenade_data[0]["id"],
            grenade_data[1]["id"],
            grenade_data[2]["id"],
        ]
    )
    ws.append(
        [
            "烟幕干扰弹投放点的x坐标 (m)",
            grenade_data[0]["p_drop"][0],
            grenade_data[1]["p_drop"][0],
            grenade_data[2]["p_drop"][0],
        ]
    )
    ws.append(
        [
            "烟幕干扰弹投放点的y坐标 (m)",
            grenade_data[0]["p_drop"][1],
            grenade_data[1]["p_drop"][1],
            grenade_data[2]["p_drop"][1],
        ]
    )
    ws.append(
        [
            "烟幕干扰弹投放点的z坐标 (m)",
            grenade_data[0]["p_drop"][2],
            grenade_data[1]["p_drop"][2],
            grenade_data[2]["p_drop"][2],
        ]
    )
    ws.append(
        [
            "烟幕干扰弹起爆点的x坐标 (m)",
            grenade_data[0]["p_detonate"][0],
            grenade_data[1]["p_detonate"][0],
            grenade_data[2]["p_detonate"][0],
        ]
    )
    ws.append(
        [
            "烟幕干扰弹起爆点的y坐标 (m)",
            grenade_data[0]["p_detonate"][1],
            grenade_data[1]["p_detonate"][1],
            grenade_data[2]["p_detonate"][1],
        ]
    )
    ws.append(
        [
            "烟幕干扰弹起爆点的z坐标 (m)",
            grenade_data[0]["p_detonate"][2],
            grenade_data[1]["p_detonate"][2],
            grenade_data[2]["p_detonate"][2],
        ]
    )
    ws.append(["有效干扰时长 (s)", max_time, "", ""])

    # 空一行
    ws.append([])

    # 写入注释
    ws.append(["注：以x轴为正向，逆时针方向为正，取值0~360（度）。"])

    # --- 3. 设置格式 ---
    # 合并单元格
    ws.merge_cells("B11:D11")  # 合并有效时长的值单元格
    ws.merge_cells("A13:D13")  # 合并注释单元格

    # 调整列宽
    ws.column_dimensions["A"].width = 40
    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 20

    # 设置数值格式
    for row in ws.iter_rows(min_row=2, max_row=11):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"

    # --- 4. 保存文件 ---
    filename = "result1.xlsx"
    try:
        wb.save(filename)
        print(f"结果已成功保存到 '{os.path.abspath(filename)}'")
    except Exception as e:
        print(f"保存文件失败: {e}")


# --- 主程序入口 ---
if __name__ == "__main__":
    # run_optimization_Q3()
    # v_FY1, theta_rad, t_drop1, t_fuse1, dt_drop2, t_fuse2, dt_drop3, t_fuse3
    ret = objective_function_Q3((70, 176.5 / 180 * np.pi, 0, 2.5, 1.2, 3.0, 1.3, 3.5))
    print(ret)
    pass
